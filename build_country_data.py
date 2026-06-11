#!/usr/bin/env python3
"""
Build country-specific career map JSON data files for 6 countries:
India (Level 3 from PLFS); Nigeria, Indonesia, Russia, Pakistan, Kenya
(Level 2 from ILOSTAT EMP_TEMP_SEX_OC2_NB, national LFS micro-data).
"""
import xml.etree.ElementTree as ET
import csv as csv_mod
import json, os, math
import openpyxl

# ─── ISCO-08 Occupation Names (Chinese + English) ───────────────────────

ISCO_L1_NAMES = {
    '1': ('管理人员', 'Managers'),
    '2': ('专业人员', 'Professionals'),
    '3': ('技术员和助理专业人员', 'Technicians & Associate Professionals'),
    '4': ('文职支持人员', 'Clerical Support Workers'),
    '5': ('服务和销售人员', 'Service & Sales Workers'),
    '6': ('农林渔业技术工人', 'Skilled Agricultural/Forestry/Fishery Workers'),
    '7': ('手工艺及相关行业工人', 'Craft & Related Trades Workers'),
    '8': ('工厂操作员和装配工', 'Plant/Machine Operators & Assemblers'),
    '9': ('基层职业', 'Elementary Occupations'),
    '0': ('军人', 'Armed Forces Occupations'),
}

ISCO_L2_NAMES = {
    '01': ('军官', 'Commissioned Armed Forces Officers'),
    '02': ('军士', 'Non-commissioned Armed Forces Officers'),
    '03': ('其他军人', 'Armed Forces Occupations, Other Ranks'),
    '11': ('首席执行官/高级官员/立法者', 'Chief Executives/Senior Officials/Legislators'),
    '12': ('行政和商务经理', 'Administrative & Commercial Managers'),
    '13': ('生产和专业服务经理', 'Production & Specialized Services Managers'),
    '14': ('酒店/零售及其他服务经理', 'Hospitality/Retail/Other Services Managers'),
    '21': ('科学和工程专业人员', 'Science & Engineering Professionals'),
    '22': ('卫生专业人员', 'Health Professionals'),
    '23': ('教学专业人员', 'Teaching Professionals'),
    '24': ('商业和行政专业人员', 'Business & Administration Professionals'),
    '25': ('信息通信技术专业人员', 'ICT Professionals'),
    '26': ('法律/社会/文化专业人员', 'Legal/Social/Cultural Professionals'),
    '31': ('科学和工程助理专业人员', 'Science & Engineering Associate Professionals'),
    '32': ('卫生助理专业人员', 'Health Associate Professionals'),
    '33': ('商业和行政助理专业人员', 'Business & Administration Associate Professionals'),
    '34': ('法律/社会/文化助理专业人员', 'Legal/Social/Cultural Associate Professionals'),
    '35': ('信息通信技术技术员', 'ICT Technicians'),
    '41': ('一般文职人员', 'General & Keyboard Clerks'),
    '42': ('客户服务人员', 'Customer Services Clerks'),
    '43': ('数字记录和材料记录人员', 'Numerical & Material Recording Clerks'),
    '44': ('其他文职支持人员', 'Other Clerical Support Workers'),
    '51': ('个人服务人员', 'Personal Service Workers'),
    '52': ('销售人员', 'Sales Workers'),
    '53': ('个人护理人员', 'Personal Care Workers'),
    '54': ('保安服务人员', 'Protective Services Workers'),
    '61': ('面向市场的农业技术工人', 'Market-oriented Skilled Agricultural Workers'),
    '62': ('面向市场的林渔猎技术工人', 'Market-oriented Skilled Forestry/Fishery/Hunting Workers'),
    '63': ('自给型农林渔业劳动者', 'Subsistence Farmers/Fishers/Hunters/Gatherers'),
    '71': ('建筑和相关行业工人', 'Building & Related Trades Workers'),
    '72': ('金属/机械及相关行业工人', 'Metal/Machinery & Related Trades Workers'),
    '73': ('手工艺和印刷工人', 'Handicraft & Printing Workers'),
    '74': ('电气和电子行业工人', 'Electrical & Electronic Trades Workers'),
    '75': ('食品加工/木工/服装及相关工人', 'Food Processing/Woodworking/Garment Workers'),
    '81': ('固定设备和机器操作员', 'Stationary Plant & Machine Operators'),
    '82': ('装配工', 'Assemblers'),
    '83': ('驾驶员和移动设备操作员', 'Drivers & Mobile Plant Operators'),
    '91': ('清洁工和帮佣', 'Cleaners & Helpers'),
    '92': ('农林渔业劳工', 'Agricultural/Forestry/Fishery Labourers'),
    '93': ('采矿/建筑/制造/运输劳工', 'Mining/Construction/Manufacturing/Transport Labourers'),
    '94': ('食品加工助理', 'Food Preparation Assistants'),
    '95': ('街头和相关销售服务人员', 'Street & Related Sales & Service Workers'),
    '96': ('垃圾收集和其他基层工人', 'Refuse Workers & Other Elementary Workers'),
}

# India NCO-2015 3-digit names (≈ ISCO-08 minor groups)
ISCO_L3_NAMES = {
    '111': ('立法者和高级官员', 'Legislators & Senior Officials'), '112': ('总经理和执行董事', 'Managing Directors & Chief Executives'),
    '121': ('商务服务和行政经理', 'Business Services & Administration Managers'), '122': ('销售/市场/开发经理', 'Sales/Marketing/Development Managers'),
    '131': ('农林渔业生产经理', 'Production Managers in Agriculture/Forestry/Fisheries'), '132': ('制造/采矿/建筑/物流经理', 'Manufacturing/Mining/Construction/Distribution Managers'),
    '133': ('信息通信技术服务经理', 'ICT Service Managers'), '134': ('专业服务经理', 'Professional Services Managers'),
    '141': ('酒店和餐饮经理', 'Hotel & Restaurant Managers'), '142': ('零售和批发贸易经理', 'Retail & Wholesale Trade Managers'),
    '143': ('其他服务经理', 'Other Services Managers'),
    '211': ('物理和地球科学专业人员', 'Physical & Earth Science Professionals'), '212': ('数学/精算/统计专业人员', 'Mathematicians/Actuaries/Statisticians'),
    '213': ('生命科学专业人员', 'Life Science Professionals'), '214': ('工程专业人员（不含电气电子）', 'Engineering Professionals (excl. Electrotechnology)'),
    '215': ('电气工程专业人员', 'Electrotechnology Engineers'), '216': ('建筑师/规划师/测量师/设计师', 'Architects/Planners/Surveyors/Designers'),
    '221': ('医生', 'Medical Doctors'), '222': ('护理和助产专业人员', 'Nursing & Midwifery Professionals'),
    '223': ('传统医学和替代医学专业人员', 'Traditional & Complementary Medicine Professionals'), '224': ('急救医师和护理人员', 'Paramedical Practitioners'),
    '225': ('兽医', 'Veterinarians'), '226': ('其他卫生专业人员', 'Other Health Professionals'),
    '231': ('大学和高等教育教师', 'University & Higher Education Teachers'), '232': ('职业教育教师', 'Vocational Education Teachers'),
    '233': ('中学教师', 'Secondary Education Teachers'), '234': ('小学和学前教师', 'Primary School & Early Childhood Teachers'),
    '235': ('其他教学专业人员', 'Other Teaching Professionals'),
    '241': ('财务专业人员', 'Finance Professionals'), '242': ('行政专业人员', 'Administration Professionals'),
    '243': ('销售/市场/公关专业人员', 'Sales/Marketing/Public Relations Professionals'),
    '251': ('软件和应用开发人员/分析师', 'Software & Applications Developers & Analysts'), '252': ('数据库和网络专业人员', 'Database & Network Professionals'),
    '261': ('法律专业人员', 'Legal Professionals'), '262': ('图书馆和档案专业人员', 'Librarians & Related Information Professionals'),
    '263': ('社会和宗教专业人员', 'Social & Religious Professionals'), '264': ('作者/记者/语言学家', 'Authors/Journalists/Linguists'),
    '265': ('创意和表演艺术家', 'Creative & Performing Artists'),
    '311': ('物理和工程科学技术员', 'Physical & Engineering Science Technicians'), '312': ('采矿/制造/建筑监督员', 'Mining/Manufacturing/Construction Supervisors'),
    '313': ('过程控制技术员', 'Process Control Technicians'), '314': ('生命科学技术员和相关助理', 'Life Science Technicians & Related Associates'),
    '315': ('船舶和航空控制及技术员', 'Ship & Aircraft Controllers & Technicians'),
    '321': ('医学和药学技术员', 'Medical & Pharmaceutical Technicians'), '322': ('护理和助产助理专业人员', 'Nursing & Midwifery Associate Professionals'),
    '323': ('传统医学和替代医学助理', 'Traditional & Complementary Medicine Associate Professionals'), '324': ('兽医技术员和助理', 'Veterinary Technicians & Assistants'),
    '325': ('其他卫生助理专业人员', 'Other Health Associate Professionals'),
    '331': ('金融和数学助理专业人员', 'Financial & Mathematical Associate Professionals'), '332': ('销售和采购代理/经纪人', 'Sales & Purchasing Agents & Brokers'),
    '333': ('商业服务代理', 'Business Services Agents'), '334': ('行政和专业秘书', 'Administrative & Specialized Secretaries'),
    '335': ('政府监管助理专业人员', 'Government Regulatory Associate Professionals'),
    '341': ('法律/社会助理专业人员', 'Legal/Social Associate Professionals'), '342': ('体育和健身工作者', 'Sports & Fitness Workers'),
    '343': ('艺术/文化/烹饪助理专业人员', 'Artistic/Cultural/Culinary Associate Professionals'),
    '351': ('信息通信技术运营技术员', 'ICT Operations Technicians'), '352': ('通信技术员', 'Telecommunications Technicians'),
    '411': ('一般办事员', 'General Office Clerks'), '412': ('秘书（一般）', 'Secretaries (General)'),
    '413': ('键盘操作员', 'Keyboard Operators'),
    '421': ('出纳员和相关人员', 'Tellers/Money Collectors & Related Clerks'), '422': ('客户信息工作者', 'Client Information Workers'),
    '431': ('数字记录和会计人员', 'Numerical Clerks'), '432': ('材料记录和运输人员', 'Material Recording & Transport Clerks'),
    '441': ('其他文职支持人员', 'Other Clerical Support Workers'),
    '511': ('旅行服务员/导游/厨师', 'Travel Attendants/Conductors/Guides'), '512': ('厨师', 'Cooks'),
    '513': ('服务员和酒吧侍应', 'Waiters & Bartenders'), '514': ('理发师/美容师及相关人员', 'Hairdressers/Beauticians & Related Workers'),
    '515': ('建筑和家务管理人员', 'Building & Housekeeping Supervisors'), '516': ('其他个人服务人员', 'Other Personal Services Workers'),
    '521': ('街头和市场销售人员', 'Street & Market Salespersons'), '522': ('商店销售人员', 'Shop Salespersons'),
    '523': ('收银员和售票员', 'Cashiers & Ticket Clerks'), '524': ('其他销售人员', 'Other Sales Workers'),
    '531': ('儿童看护人员', 'Child Care Workers'), '532': ('个人护理人员（卫生机构）', 'Personal Care Workers in Health Services'),
    '541': ('保安人员', 'Protective Services Workers'),
    '611': ('市场园艺师和作物种植者', 'Market Gardeners & Crop Growers'), '612': ('畜牧工人', 'Animal Producers'),
    '613': ('混合作物和畜牧农民', 'Mixed Crop & Animal Producers'),
    '621': ('林业和相关工人', 'Forestry & Related Workers'), '622': ('渔业工人/猎人/捕猎者', 'Fishery Workers/Hunters/Trappers'),
    '631': ('自给型作物种植者', 'Subsistence Crop Farmers'), '632': ('自给型畜牧者', 'Subsistence Livestock Farmers'),
    '633': ('自给型混合种植和畜牧者', 'Subsistence Mixed Crop & Livestock Farmers'), '634': ('自给型渔民/猎人/采集者', 'Subsistence Fishers/Hunters/Gatherers'),
    '711': ('建筑工人', 'House Builders'), '712': ('建筑装修和相关工人', 'Building Finishers & Related Trades Workers'),
    '713': ('油漆工/建筑物清洁工及相关', 'Painters/Building Structure Cleaners & Related'),
    '721': ('钣金工/结构金属工及相关', 'Sheet & Structural Metal Workers & Related'), '722': ('铁匠/锻工和相关工具制造工', 'Blacksmiths/Toolmakers & Related'),
    '723': ('机械装配工和修理工', 'Machinery Mechanics & Repairers'),
    '731': ('手工艺工人', 'Handicraft Workers'), '732': ('印刷行业工人', 'Printing Trades Workers'),
    '741': ('电气设备安装和修理工', 'Electrical Equipment Installers & Repairers'), '742': ('电子和通信安装/修理工', 'Electronics & Telecommunications Installers'),
    '751': ('食品加工及相关行业工人', 'Food Processing & Related Trades Workers'), '752': ('木材加工工人/家具制造工', 'Wood Treaters/Cabinet-makers'),
    '753': ('服装及相关行业工人', 'Garment & Related Trades Workers'), '754': ('其他手工艺和相关行业工人', 'Other Craft & Related Workers'),
    '811': ('采矿和矿物加工设备操作员', 'Mining & Mineral Processing Plant Operators'), '812': ('金属加工和精加工设备操作员', 'Metal Processing & Finishing Plant Operators'),
    '813': ('化工和照相制品设备操作员', 'Chemical & Photographic Products Plant Operators'), '814': ('橡胶/塑料/纸张产品机器操作员', 'Rubber/Plastic/Paper Products Machine Operators'),
    '815': ('纺织/皮革/毛皮产品机器操作员', 'Textile/Fur/Leather Products Machine Operators'), '816': ('食品和相关产品机器操作员', 'Food & Related Products Machine Operators'),
    '817': ('木材加工和造纸机器操作员', 'Wood Processing & Papermaking Plant Operators'), '818': ('其他固定设备和机器操作员', 'Other Stationary Plant & Machine Operators'),
    '821': ('装配工', 'Assemblers'),
    '831': ('火车司机和相关人员', 'Locomotive Engine Drivers & Related'), '832': ('汽车/货车/摩托车驾驶员', 'Car/Van/Motorcycle Drivers'),
    '833': ('重型卡车和客车驾驶员', 'Heavy Truck & Bus Drivers'), '834': ('移动设备操作员', 'Mobile Plant Operators'),
    '835': ('船舶和飞机控制/操作人员', 'Ships Deck Crews & Related'),
    '911': ('家庭佣工和帮手', 'Domestic/Hotel/Office Cleaners & Helpers'), '912': ('车辆/窗户/洗衣及其他手工清洁工', 'Vehicle/Window/Laundry & Other Hand Cleaning Workers'),
    '921': ('农林渔业劳工', 'Agricultural/Forestry/Fishery Labourers'),
    '931': ('采矿和建筑劳工', 'Mining & Construction Labourers'), '932': ('制造业劳工', 'Manufacturing Labourers'),
    '933': ('运输和仓储劳工', 'Transport & Storage Labourers'),
    '941': ('食品加工助理', 'Food Preparation Assistants'),
    '951': ('街头和相关服务工人', 'Street & Related Service Workers'), '952': ('街头小贩（不含食品）', 'Street Vendors (excl. Food)'),
    '961': ('垃圾收集工人', 'Refuse Workers'), '962': ('其他基层工人', 'Other Elementary Workers'),
}

# ─── Education mapping per country ──────────────────────────────────────

EDUCATION_SYSTEMS = {
    'IND': {
        'name': '学历要求（印度教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('高中/12年级以下（对标：初中及以下）', 'Below Higher Secondary', 1),
            ('高中/12年级毕业（对标：高中文凭）', 'Higher Secondary (10+2)', 2),
            ('ITI/文凭课程（对标：职业教育证书）', 'ITI/Diploma', 3),
            ('学士学位（对标：学士学位）', "Bachelor's Degree (B.A./B.Sc./B.Tech.)", 4),
            ('硕士学位（对标：硕士学位）', "Master's Degree (M.A./M.Sc./M.Tech.)", 5),
            ('博士/专业学位（对标：博士学位）', 'Ph.D./Professional Degree (M.D./LL.B.)', 6),
        ]
    },
    'NGA': {
        'name': '学历要求（尼日利亚教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('小学毕业证书（对标：小学教育）', 'First School Leaving Certificate', 1),
            ('SSCE/WAEC（对标：高中文凭）', 'SSCE/WAEC (Senior Secondary)', 2),
            ('OND/NCE（对标：副学士/职业教育）', 'OND/NCE (Ordinary National Diploma)', 3),
            ('HND/学士学位（对标：学士学位）', "HND/Bachelor's Degree", 4),
            ('硕士学位（对标：硕士学位）', "Master's Degree (M.Sc./MBA)", 5),
            ('博士/专业学位（对标：博士学位）', 'Ph.D./Professional Degree', 6),
        ]
    },
    'IDN': {
        'name': '学历要求（印尼教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('SD/SMP毕业（对标：初中及以下）', 'SD/SMP (Primary/Junior Secondary)', 1),
            ('SMA/SMK毕业（对标：高中文凭）', 'SMA/SMK (Senior Secondary/Vocational)', 2),
            ('D1-D3文凭（对标：副学士/职业教育）', 'Diploma I-III (D1/D2/D3)', 3),
            ('S1学士学位（对标：学士学位）', 'Sarjana (S1)', 4),
            ('S2硕士学位（对标：硕士学位）', 'Magister (S2)', 5),
            ('S3博士学位（对标：博士学位）', 'Doktor (S3)', 6),
        ]
    },
    'RUS': {
        'name': '学历要求（俄罗斯教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('基础普通教育（对标：初中及以下）', 'Basic General Education (9 years)', 1),
            ('中等普通教育（对标：高中文凭）', 'Secondary General Education (11 years)', 2),
            ('中等职业教育（对标：职业教育证书）', 'Secondary Vocational Education (SPO)', 3),
            ('高等教育-学士（对标：学士学位）', 'Higher Education - Bakalavr (Bachelor)', 4),
            ('高等教育-硕士/专家（对标：硕士学位）', 'Higher Education - Magistr/Specialist', 5),
            ('研究生学位（对标：博士学位）', 'Kandidat/Doktor Nauk (Postgraduate)', 6),
        ]
    },
    'PAK': {
        'name': '学历要求（巴基斯坦教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('小学/初中（对标：初中及以下）', 'Primary/Middle (Class 1-8)', 1),
            ('中学/高中（对标：高中文凭）', 'Matric/Intermediate (SSC/HSSC)', 2),
            ('文凭/技术证书（对标：职业教育证书）', 'Diploma/Technical Certificate (DAE)', 3),
            ('学士学位（对标：学士学位）', "Bachelor's Degree (BA/BSc/BS)", 4),
            ('硕士学位（对标：硕士学位）', "Master's Degree (MA/MSc/MPhil)", 5),
            ('博士/专业学位（对标：博士学位）', 'Ph.D./Professional Degree (MBBS/LLB)', 6),
        ]
    },
    'KEN': {
        'name': '学历要求（肯尼亚教育体系）',
        'levels': [
            ('无正规学历（对标：无正式教育）', 'No formal education', 0),
            ('小学教育（对标：初中及以下）', 'Primary Education (KCPE)', 1),
            ('中学教育（对标：高中文凭）', 'Secondary Education (KCSE)', 2),
            ('技术职业证书/文凭（对标：职业教育证书）', 'TVET Certificate/Diploma', 3),
            ('学士学位（对标：学士学位）', "Bachelor's Degree", 4),
            ('硕士学位（对标：硕士学位）', "Master's Degree", 5),
            ('博士/专业学位（对标：博士学位）', 'Ph.D./Professional Degree', 6),
        ]
    },
}

# Map ISCO major groups to typical education level index (0-6)
ISCO_MAJOR_EDU = {
    '0': 3, '1': 4, '2': 5, '3': 3, '4': 2, '5': 2, '6': 1, '7': 2, '8': 2, '9': 1,
}
ISCO_L2_EDU_OVERRIDE = {
    '11': 5, '12': 4, '13': 4, '14': 3,
    '21': 5, '22': 6, '23': 5, '24': 4, '25': 4, '26': 5,
    '31': 3, '32': 3, '33': 3, '34': 3, '35': 3,
    '41': 2, '42': 2, '43': 2, '44': 2,
    '51': 1, '52': 1, '53': 2, '54': 2,
    '61': 1, '62': 1, '63': 0,
    '71': 2, '72': 2, '73': 2, '74': 2, '75': 1,
    '81': 1, '82': 2, '83': 1,
    '91': 0, '92': 0, '93': 0, '94': 0, '95': 0, '96': 0,
}

# AI Exposure ratings by ISCO L2 code
AI_EXPOSURE = {
    '01': 4, '02': 3, '03': 2,
    '11': 7, '12': 7, '13': 6, '14': 5,
    '21': 8, '22': 6, '23': 6, '24': 8, '25': 9, '26': 7,
    '31': 6, '32': 5, '33': 7, '34': 6, '35': 7,
    '41': 9, '42': 8, '43': 9, '44': 8,
    '51': 3, '52': 4, '53': 3, '54': 3,
    '61': 2, '62': 2, '63': 1,
    '71': 2, '72': 3, '73': 5, '74': 3, '75': 2,
    '81': 4, '82': 4, '83': 3,
    '91': 1, '92': 1, '93': 1, '94': 1, '95': 2, '96': 1,
}

AI_EXPOSURE_RATIONALE = {
    '11': '高级管理和立法工作涉及大量信息处理、决策和沟通，AI可辅助数据分析和报告生成，但最终决策仍需人类判断。',
    '12': '行政和商务管理工作高度依赖数据分析、报告和沟通，AI可显著提升效率，但团队领导和战略决策仍需人际技能。',
    '13': '生产管理需现场监督和设备相关知识，AI可优化调度和质量控制，但物理现场管理难以替代。',
    '14': '服务业管理涉及客户互动和现场管理，AI可辅助运营优化，但人际服务管理仍需人类参与。',
    '21': '科学和工程工作高度数字化，AI在建模、数据分析和设计优化方面能力很强，但创新性研究仍需人类创造力。',
    '22': '卫生专业人员的诊断和治疗正越来越多地利用AI辅助，但直接患者护理和手术等核心工作仍需人类执行。',
    '23': '教学工作正被AI辅助工具改变（个性化学习、自动批改等），但课堂管理和学生情感支持难以替代。',
    '24': '商业和行政专业工作高度数字化，AI在财务分析、合规检查和报告生成方面表现出色。',
    '25': '信息通信技术专业人员的核心工作——编程、系统设计、数据分析——正是AI能力快速提升的领域。',
    '26': '法律研究、写作和文化创作正受到生成式AI的深刻影响，但需要创造力和人类判断。',
    '31': '科学和工程技术员工作包含实验操作和设备维护等物理成分，AI影响适中。',
    '32': '卫生助理工作涉及直接患者接触，AI可辅助诊断但不能替代护理行为。',
    '33': '商业助理工作高度依赖数据处理和客户沟通，AI可显著提升效率。',
    '34': '法律和社会助理工作涉及文档处理和研究，AI可大幅加速这些任务。',
    '35': 'IT技术员负责系统维护和故障排除，AI可辅助诊断但物理操作仍需人类。',
    '41': '一般文职工作几乎完全是数字化的——数据输入、文档处理、日程安排——是AI自动化的首要目标。',
    '42': '客户服务工作正被AI聊天机器人和虚拟助手快速改变，但复杂投诉仍需人工处理。',
    '43': '数字记录和会计工作高度程式化，AI和RPA可自动化大部分任务。',
    '44': '其他文职工作涉及多种行政任务，AI可自动化其中大部分。',
    '51': '个人服务工作如厨师、服务员等需要物理操作和人际互动，AI影响较小。',
    '52': '销售工作中的推荐和分析部分可被AI增强，但面对面销售和人际关系仍是关键。',
    '53': '护理工作需要直接身体接触和情感支持，是AI最难替代的领域之一。',
    '54': '安保工作需要物理存在和实时判断，AI可辅助监控但不能替代现场人员。',
    '61': '农业工作主要在户外进行，涉及植物和土壤的物理操作，AI在精准农业方面有辅助作用。',
    '62': '林业和渔业工作在自然环境中进行，物理性强，AI影响有限。',
    '63': '自给型农业几乎完全是体力劳动，AI影响极小。',
    '71': '建筑工作需要体力劳动和现场技能，AI可辅助设计但施工难以自动化。',
    '72': '金属和机械工作需要手工技能，AI可优化流程但核心操作需人类完成。',
    '73': '手工艺和印刷工作中，AI在设计领域有显著影响，但手工制作仍需人类技能。',
    '74': '电气和电子安装需要现场操作技能，AI可辅助诊断但物理工作难以替代。',
    '75': '食品加工和服装制造中部分工作可自动化，但手工操作仍有需求。',
    '81': '固定设备操作正越来越多地通过自动化控制，AI可优化操作参数。',
    '82': '装配工作中的重复性部分易被机器人替代，但复杂装配仍需人类技能。',
    '83': '驾驶工作面临自动驾驶技术的长期影响，但短期内仍大量需要人类驾驶员。',
    '91': '清洁工作是基本体力劳动，AI影响极小。',
    '92': '农林渔业劳工从事基础体力工作，AI影响极小。',
    '93': '采矿和建筑劳工从事繁重体力工作，AI可提升安全性但不能替代人力。',
    '94': '食品加工助理从事基本体力工作，AI影响极小。',
    '95': '街头销售和服务涉及人际互动，AI影响有限。',
    '96': '其他基层工作多为体力劳动，AI影响极小。',
}

# ─── Parse functions ────────────────────────────────────────────────────

def parse_xml_series(filepath, key_dim):
    """Generic XML parser for ILOSTAT SDMX data"""
    tree = ET.parse(filepath)
    root = tree.getroot()
    results = {}
    for elem in root.iter():
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag == 'Series':
            keys = {}
            for child in elem:
                ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if ctag == 'SeriesKey':
                    for val in child:
                        keys[val.get('id','')] = val.get('value','')
            main_key = keys.get(key_dim, '')
            if not main_key:
                continue
            extra_keys = {k: v for k, v in keys.items() if k != key_dim and k != 'FREQ' and k != 'REF_AREA' and k != 'SEX' and k != 'MEASURE'}
            full_key = main_key if not extra_keys else f"{main_key}|{'|'.join(f'{k}={v}' for k,v in sorted(extra_keys.items()))}"
            for child in elem:
                ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if ctag == 'Obs':
                    year = None; value = None
                    for sub in child:
                        stag = sub.tag.split('}')[-1] if '}' in sub.tag else sub.tag
                        if stag == 'ObsDimension': year = sub.get('value','')
                        elif stag == 'ObsValue': value = sub.get('value','')
                    if year and value:
                        if full_key not in results: results[full_key] = {}
                        results[full_key][year] = float(value)
    return results

def get_latest_year_data(series_data, min_code_len=2):
    """Get the latest year's data for each occupation code"""
    result = {}
    for code, years in series_data.items():
        base_code = code.split('|')[0]
        # Extract the numeric part
        parts = base_code.split('_')
        num = parts[-1] if parts else ''
        if num in ('TOTAL', 'X') or len(num) < min_code_len:
            continue
        if years:
            latest = max(years.keys())
            result[num] = {'value': years[latest], 'year': latest, 'history': years}
    return result

def load_oc2_timeseries(country_code, sex='SEX_T'):
    """Load ISCO-08 Level-2 employment time series from ILOSTAT SDMX CSV.

    Source: https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/{CC}.A..SEX_T+SEX_M+SEX_F.
    (indicator EMP_TEMP_SEX_OC2_NB, national LFS micro-data processed by ILO).
    Values are in thousands (UNIT_MULT=3). Returns {year: {l2_code: persons}}.

    sex: which SEX dimension value to load (SEX_T=total, SEX_M=male, SEX_F=female).
         Totals and trends always use SEX_T so they stay consistent.
    """
    path = f'ilostat_data/{country_code}_oc2_timeseries.csv'
    data = {}
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            if row.get('SEX', 'SEX_T') != sex:
                continue
            oc2 = row.get('OC2', '')
            val = row.get('OBS_VALUE', '')
            if 'ISCO08' not in oc2 or not val:
                continue
            code = oc2.replace('OC2_ISCO08_', '')
            if code in ('TOTAL', 'X') or not code.isdigit() or len(code) != 2:
                continue
            year = row['TIME_PERIOD']
            mult = 10 ** int(row.get('UNIT_MULT') or 0)
            data.setdefault(year, {})[code] = float(val) * mult
    return data


def load_sex_breakdown(country_code, year):
    """Per-occupation (ISCO L2) male/female employment for a given survey year.

    Returns {l2_code: {'male': persons, 'female': persons}} using the same
    ILOSTAT EMP_TEMP_SEX_OC2_NB source as the totals (so figures are consistent).
    Only includes occupations where both SEX_M and SEX_F are available.
    """
    male = load_oc2_timeseries(country_code, sex='SEX_M').get(year, {})
    female = load_oc2_timeseries(country_code, sex='SEX_F').get(year, {})
    out = {}
    for code in set(male) | set(female):
        m = male.get(code)
        f = female.get(code)
        if m is None or f is None:
            continue
        out[code] = {'male': m, 'female': f}
    return out


def sex_fields(breakdown, l2_code):
    """Build the male/female/female_share output fields for an occupation.

    breakdown: dict from load_sex_breakdown
    l2_code:   2-digit ISCO L2 code (L3 occupations inherit their L2 group)
    Returns a dict to merge into the occupation record. female_share is the
    female fraction of (male+female) employment, 0-100, rounded to 1 decimal;
    None when no data.
    """
    b = breakdown.get(l2_code)
    if not b:
        return {'male': None, 'female': None, 'female_share': None}
    m, f = b['male'], b['female']
    tot = m + f
    return {
        'male': int(m),
        'female': int(f),
        'female_share': round(f / tot * 100, 1) if tot > 0 else None,
    }


def compute_l2_share_changes(ts, year_from, year_to):
    """Per-occupation (ISCO L2) employment share change in percentage points,
    computed from actual LFS survey data (not modelled estimates)."""
    a, b = ts.get(year_from, {}), ts.get(year_to, {})
    ta, tb = sum(a.values()), sum(b.values())
    changes = {}
    if ta <= 0 or tb <= 0:
        return changes
    for code, v in b.items():
        if code in a:
            changes[code] = {
                'change': round(v / tb * 100 - a[code] / ta * 100, 2),
                'year_from': year_from,
                'year_to': year_to,
            }
    return changes


def compute_share_change(hist_data, years_back=5):
    """Compute employment share change over specified years"""
    # Get all occupation codes and their time series
    total_by_year = {}
    occ_by_year = {}
    
    for code, years in hist_data.items():
        for yr, val in years.items():
            if yr not in total_by_year:
                total_by_year[yr] = 0
                occ_by_year[yr] = {}
            total_by_year[yr] += val
            occ_by_year[yr][code] = val
    
    sorted_years = sorted(total_by_year.keys())
    if len(sorted_years) < 2:
        return {}
    
    latest_year = sorted_years[-1]
    earlier_year = None
    for yr in sorted_years:
        if int(latest_year) - int(yr) >= years_back:
            earlier_year = yr
    if not earlier_year:
        earlier_year = sorted_years[0]
    
    changes = {}
    for code in occ_by_year.get(latest_year, {}):
        if code in occ_by_year.get(earlier_year, {}) and total_by_year.get(latest_year, 0) > 0 and total_by_year.get(earlier_year, 0) > 0:
            share_now = occ_by_year[latest_year][code] / total_by_year[latest_year] * 100
            share_then = occ_by_year[earlier_year][code] / total_by_year[earlier_year] * 100
            changes[code] = {
                'share_now': round(share_now, 2),
                'share_then': round(share_then, 2),
                'change': round(share_now - share_then, 2),
                'year_from': earlier_year,
                'year_to': latest_year,
            }
    return changes

# ─── Build India data (Level 3 from PLFS) ───────────────────────────────

def build_india_data():
    """Build India data using PLFS Table 25 (3-digit NCO) + ILOSTAT Level 2 for shares"""
    print("Building India data...")

    # Real per-occupation (ISCO L2) share changes from PLFS survey data 2022-2025
    # (replaces the previous ILO modelled L1 estimates which produced implausible
    #  artifacts like -12.55pp for managers)
    ts = load_oc2_timeseries('IND')
    ts_years = sorted(ts.keys())
    share_changes_l2 = compute_l2_share_changes(ts, ts_years[0], ts_years[-1])

    # Per-occupation male/female breakdown (ISCO L2) from latest ILOSTAT year;
    # each 3-digit PLFS occupation inherits its 2-digit group's sex split.
    sex_bd = load_sex_breakdown('IND', ts_years[-1])
    
    # Parse PLFS Table 25 - percentage distributions at 3-digit level
    wb = openpyxl.load_workbook('ilostat_data/India_PLFS_Table25.xlsx')
    ws = wb['Sheet1']
    
    occupations = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if code.isdigit() and len(code) == 3:
            # Column 8 (index 7) = rural+urban person percentage, column 9 = rural+urban male/female
            pct_person = row[7]  # rural+urban, person
            if pct_person and isinstance(pct_person, (int, float)) and pct_person > 0:
                occupations.append({
                    'code': code,
                    'pct': float(pct_person),
                })
    
    # Total employment India 2024: ~483 million (from ILOSTAT)
    total_employment = 482793235
    
    # Build occupation list
    result = []
    for occ in occupations:
        code = occ['code']
        major = code[0]
        l2_code = code[:2]
        
        # Get name
        name_cn, name_en = ISCO_L3_NAMES.get(code, (f'职业代码 {code}', f'Occupation {code}'))
        
        # Compute jobs from percentage
        jobs = int(total_employment * occ['pct'] / 100)
        
        # Category = major group
        cat_cn, cat_en = ISCO_L1_NAMES.get(major, (f'大类{major}', f'Major Group {major}'))
        
        # Employment share change: real survey data at L2 (L3 history not available),
        # each 3-digit occupation inherits its 2-digit sub-major group's change
        share_info = share_changes_l2.get(l2_code, {})
        share_change = share_info.get('change', None)
        
        # Education
        edu_idx = ISCO_L2_EDU_OVERRIDE.get(l2_code, ISCO_MAJOR_EDU.get(major, 2))
        edu_label = EDUCATION_SYSTEMS['IND']['levels'][min(edu_idx, 6)][0]
        
        # AI exposure
        ai_exp = AI_EXPOSURE.get(l2_code, AI_EXPOSURE.get(f'0{major}' if len(major)==1 else major, 5))
        ai_rationale = AI_EXPOSURE_RATIONALE.get(l2_code, '')
        
        occ = {
            'title': name_cn,
            'title_en': name_en,
            'code': code,
            'category': f'{major}-{cat_cn}',
            'jobs': jobs,
            'share_change': share_change,
            'share_change_desc': f"{share_info.get('year_from','')}-{share_info.get('year_to','')}" if share_change is not None else '',
            'education': edu_label,
            'education_idx': edu_idx,
            'exposure': ai_exp,
            'exposure_rationale': ai_rationale,
        }
        occ.update(sex_fields(sex_bd, l2_code))
        result.append(occ)
    
    return result

# ─── Build other countries data (Level 2 from ILOSTAT) ──────────────────

def build_country_l2(country_code, level_year=None, share_years=None):
    """Build data for Nigeria, Indonesia, Russia using ILOSTAT ISCO-08 Level 2.

    level_year:  which survey year to use for employment levels (default: latest)
    share_years: (year_from, year_to) to compute real per-occupation share changes
                 from LFS survey data. If None, falls back to ILO modelled
                 estimates at the major-group (L1) level.
    """
    print(f"Building {country_code} data...")

    ts = load_oc2_timeseries(country_code)
    ts_years = sorted(ts.keys())
    year = level_year or ts_years[-1]
    levels = ts[year]

    # Per-occupation male/female breakdown for the same survey year
    sex_bd = load_sex_breakdown(country_code, year)

    # Per-occupation share changes from real survey data when possible
    share_changes_l2 = {}
    share_changes_l1 = {}
    if share_years:
        share_changes_l2 = compute_l2_share_changes(ts, share_years[0], share_years[1])
    else:
        # Fallback: ILO modelled estimates, major group (L1) only
        hist_data = parse_xml_series(f'ilostat_data/{country_code}_model_hist.xml', 'OCU')
        hist_clean = {}
        for k, v in hist_data.items():
            base = k.split('|')[0]
            num = base.split('_')[-1]
            if num.isdigit() and len(num) == 1:
                hist_clean[num] = v
        share_changes_l1 = compute_share_change(hist_clean)

    result = []
    for num_code, value in sorted(levels.items()):
        major = num_code[0]
        name_cn, name_en = ISCO_L2_NAMES.get(num_code, (f'职业代码 {num_code}', f'Occupation {num_code}'))
        cat_cn, cat_en = ISCO_L1_NAMES.get(major, (f'大类{major}', f'Major Group {major}'))

        jobs = int(value)

        # Share change: prefer real L2 survey data, else modelled L1
        share_info = share_changes_l2.get(num_code) or share_changes_l1.get(major, {})
        share_change = share_info.get('change', None)
        
        # Education
        edu_idx = ISCO_L2_EDU_OVERRIDE.get(num_code, ISCO_MAJOR_EDU.get(major, 2))
        edu_label = EDUCATION_SYSTEMS[country_code]['levels'][min(edu_idx, 6)][0]
        
        # AI exposure
        ai_exp = AI_EXPOSURE.get(num_code, 5)
        ai_rationale = AI_EXPOSURE_RATIONALE.get(num_code, '')
        
        occ = {
            'title': name_cn,
            'title_en': name_en,
            'code': num_code,
            'category': f'{major}-{cat_cn}',
            'jobs': jobs,
            'share_change': share_change,
            'share_change_desc': f"{share_info.get('year_from','')}-{share_info.get('year_to','')}" if share_change is not None else '',
            'education': edu_label,
            'education_idx': edu_idx,
            'exposure': ai_exp,
            'exposure_rationale': ai_rationale,
        }
        occ.update(sex_fields(sex_bd, num_code))
        result.append(occ)
    
    return result

# ─── Country descriptions ───────────────────────────────────────────────

COUNTRY_DESCRIPTIONS = {
    'IND': {
        'title': '印度职业地图',
        'description': '本工具可视化展示了印度<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为印度统计和计划执行部（MOSPI）发布的《定期劳动力调查》（PLFS 2023-24）和国际劳工组织（ILO）统计数据。职业分类采用NCO-2015标准（对应ISCO-08的3位小类级别）；就业趋势为PLFS调查微观数据2022→2025年各亚大类真实份额变化。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：印度MOSPI PLFS 2023-24 / ILO ILOSTAT',
        'source_links': [
            {'label': '印度MOSPI《定期劳动力调查年报 PLFS 2023-24》', 'url': 'https://www.mospi.gov.in/publication/annual-report-plfs-2023-24'},
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，印度）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=IND'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/IND.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2024',
    },
    'NGA': {
        'title': '尼日利亚职业地图',
        'description': '本工具可视化展示了尼日利亚<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为国际劳工组织（ILO）ILOSTAT数据库，基于尼日利亚国家统计局（NBS）新版全国劳动力调查（NLFS 2023，职业分类覆盖率约99.8%；2024年波次约23%就业未分类，故不采用）。职业分类采用ISCO-08标准（2位亚大类级别）；就业趋势采用ILO建模估计（大类级别，2020-2025）。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：ILO ILOSTAT / 尼日利亚NBS劳动力调查 2023',
        'source_links': [
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，尼日利亚）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=NGA'},
            {'label': '尼日利亚NBS 劳动力调查报告库', 'url': 'https://nigerianstat.gov.ng/elibrary?queries=labour%20force'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/NGA.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2023',
    },
    'IDN': {
        'title': '印度尼西亚职业地图',
        'description': '本工具可视化展示了印度尼西亚<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为国际劳工组织（ILO）ILOSTAT数据库，基于印尼中央统计局（BPS）Sakernas全国劳动力调查。职业分类采用ISCO-08标准（2位亚大类级别）；因ISCO-08口径仅有2023年一期调查数据，就业趋势采用ILO建模估计（大类级别，2020-2025）。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：ILO ILOSTAT / 印尼BPS Sakernas 2023',
        'source_links': [
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，印尼）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=IDN'},
            {'label': '印尼BPS 劳动力统计专题', 'url': 'https://www.bps.go.id/en/statistics-table?subject=520'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/IDN.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2023',
    },
    'RUS': {
        'title': '俄罗斯职业地图',
        'description': '本工具可视化展示了俄罗斯<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为国际劳工组织（ILO）ILOSTAT数据库，基于俄罗斯联邦统计局（Rosstat）劳动力调查（2025年）。职业分类采用ISCO-08标准（2位亚大类级别）；就业趋势为Rosstat劳动力调查数据2020→2025年各职业真实份额变化。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：ILO ILOSTAT / 俄罗斯Rosstat劳动力调查 2025',
        'source_links': [
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，俄罗斯）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=RUS'},
            {'label': '俄罗斯Rosstat 劳动力调查专题（俄文）', 'url': 'https://rosstat.gov.ru/labour_force'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/RUS.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2025',
    },
    'PAK': {
        'title': '巴基斯坦职业地图',
        'description': '本工具可视化展示了巴基斯坦<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为国际劳工组织（ILO）ILOSTAT数据库，基于巴基斯坦统计局（PBS）《劳动力调查》（Labour Force Survey）。职业分类采用ISCO-08标准（2位亚大类级别），2025年波次职业未分类比例极低（约0.04%）；就业趋势为劳动力调查微观数据2020→2025年各职业真实份额变化。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：ILO ILOSTAT / 巴基斯坦PBS劳动力调查 2025',
        'source_links': [
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，巴基斯坦）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=PAK'},
            {'label': '巴基斯坦统计局（PBS）劳动力调查', 'url': 'https://www.pbs.gov.pk/lfs-publications'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/PAK.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2025',
    },
    'KEN': {
        'title': '肯尼亚职业地图',
        'description': '本工具可视化展示了肯尼亚<b>{occ_count}个职业分类</b>的就业分布数据，涵盖约<b>{total_jobs_display}</b>就业人口。数据来源为国际劳工组织（ILO）ILOSTAT数据库，基于肯尼亚国家统计局（KNBS）连续住户调查（Continuous Household Survey）。职业分类采用ISCO-08标准（2位亚大类级别）；就业趋势为连续住户调查微观数据2021→2022年各职业真实份额变化。每个方块的<b>面积</b>与就业人数成正比，<b>颜色</b>展示所选指标。',
        'source': '数据来源：ILO ILOSTAT / 肯尼亚KNBS连续住户调查 2022',
        'source_links': [
            {'label': 'ILO ILOSTAT 就业人数×职业（ISCO-08 2位，肯尼亚）', 'url': 'https://rshiny.ilo.org/dataexplorer56/?lang=en&id=EMP_TEMP_SEX_OC2_NB_A&ref_area=KEN'},
            {'label': '肯尼亚国家统计局（KNBS）报告库', 'url': 'https://www.knbs.or.ke/all-reports/'},
            {'label': 'ILOSTAT SDMX API 原始数据（CSV）', 'url': 'https://sdmx.ilo.org/rest/data/ILO,DF_EMP_TEMP_SEX_OC2_NB/KEN.A..SEX_T.?format=csv&startPeriod=2015'},
        ],
        'data_year': '2022',
    },
}

# ─── Main build ─────────────────────────────────────────────────────────

def format_jobs_display(n):
    if n >= 1e8:
        return f"{n/1e8:.1f}亿"
    elif n >= 1e4:
        return f"{n/1e4:.0f}万"
    else:
        return f"{n:,.0f}"

os.makedirs('site', exist_ok=True)

all_country_data = {}

# Build India (Level 3)
india_data = build_india_data()
all_country_data['IND'] = india_data
print(f"  India: {len(india_data)} occupations, {sum(d['jobs'] for d in india_data):,.0f} total jobs")

# Build other 3 countries (Level 2)
# Russia: latest survey year (2025), real per-occupation share changes 2020->2025
# Nigeria: use 2023 NLFS levels — the 2024 wave reports ~23% of employment as
#          "occupation not elsewhere classified" (21.7M of 93.1M), which would
#          badly distort the occupation treemap; 2023 has near-complete coverage
#          (~0.2% unclassified). Share-change kept on ILO modelled L1 estimates.
# Indonesia: only 2023 available under ISCO-08, fall back to ILO modelled L1 trend
# Pakistan: latest LFS wave (2025), occupation classification near-complete
#           (~0.04% unclassified), real per-occupation share changes 2020->2025
# Kenya: Continuous Household Survey; ISCO-08 available for 2019/2021/2022.
#        Use 2022 levels; real per-occupation share changes 2021->2022 (same
#        survey instrument, comparable coverage). 2019 wave used a different
#        coverage and is not directly comparable for trend computation.
L2_BUILD_PARAMS = {
    'NGA': {'level_year': '2023', 'share_years': None},
    'IDN': {'level_year': '2023', 'share_years': None},
    'RUS': {'level_year': '2025', 'share_years': ('2020', '2025')},
    'PAK': {'level_year': '2025', 'share_years': ('2020', '2025')},
    'KEN': {'level_year': '2022', 'share_years': ('2021', '2022')},
}
for cc in ['NGA', 'IDN', 'RUS', 'PAK', 'KEN']:
    country_data = build_country_l2(cc, **L2_BUILD_PARAMS[cc])
    all_country_data[cc] = country_data
    print(f"  {cc}: {len(country_data)} occupations, {sum(d['jobs'] for d in country_data):,.0f} total jobs")

# Write combined data file
output = {
    'countries': {},
    'education_systems': EDUCATION_SYSTEMS,
}

for cc, occupations in all_country_data.items():
    total_jobs = sum(d['jobs'] for d in occupations)
    desc = COUNTRY_DESCRIPTIONS[cc]
    output['countries'][cc] = {
        'name_cn': {'IND': '印度', 'NGA': '尼日利亚', 'IDN': '印度尼西亚', 'RUS': '俄罗斯', 'PAK': '巴基斯坦', 'KEN': '肯尼亚'}[cc],
        'name_en': {'IND': 'India', 'NGA': 'Nigeria', 'IDN': 'Indonesia', 'RUS': 'Russia', 'PAK': 'Pakistan', 'KEN': 'Kenya'}[cc],
        'title': desc['title'],
        'description': desc['description'].format(
            occ_count=len(occupations),
            total_jobs_display=format_jobs_display(total_jobs),
        ),
        'source': desc['source'],
        'source_links': desc.get('source_links', []),
        'data_year': desc['data_year'],
        'total_jobs': total_jobs,
        'occ_count': len(occupations),
        'education_system': EDUCATION_SYSTEMS[cc],
        'occupations': occupations,
    }

with open('site/data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=1)

print(f"\nWrote site/data.json ({os.path.getsize('site/data.json'):,} bytes)")
print("Done!")
